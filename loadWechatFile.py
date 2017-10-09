import itchat
from itchat.content import TEXT, PICTURE, FRIENDS, CARD, MAP, SHARING, RECORDING, ATTACHMENT, VIDEO
import re
import os
import datetime
import openpyxl
import collections

def sendMessageToFriend(remark_name, content_):
    """
    :param remark_name: str obj. 可以传入用户原来的微信名 或者 你标注的名称
    :param content_: str obj.
    :return:

    nickName: 是微信用户自己取的名字
    remarkName: 是标注(别人)的名字
    """
    try:
        author = itchat.search_friends(nickName=remark_name)[0]
    except IndexError as e:
        print(e)
        try:
            author = itchat.search_friends(remarkName=remark_name)[0]
        except IndexError as e:
            print(e)
    try:
        author.send(content_)
    except UnboundLocalError as e:
        print(e)
# sendMessageToFriend('岳雪婷', "你在和机器人对话")


def sendMessageToChatroom(remark_name, content_):
    """
    :param remark_name: str obj. 只能传入群聊原来的名称
    :param content_: str obj.
    :return:
    """
    try:
        chatroom = itchat.search_chatrooms(name=remark_name)[0]
    except IndexError as e:
        print(e)
    try:
        chatroom.send(content_)
    except UnboundLocalError as e:
        print(e)
# sendMessageToChatroom('饭醉与赌博小分队', "你在和机器人对话")
    

def processMsg(msg, msg_from):
    """
    :param msg:
    :param msg_from:
    :return:
    """
    global current_date
    global rejected_chars
    if msg['FileName'] == "新邮件通知":
        msg_content = "Email from {}".format(msg['User']['NickName'])
    elif msg['Type'] == 'Text' or msg_from == 'weixin':     
        msg_content = msg['Text']     
    #如果发送的消息是附件、视屏、图片、语音, 分享
    elif msg['Type'] == "Attachment" or msg['Type'] == "Video" or msg['Type'] == 'Picture' or msg['Type'] == 'Recording' or msg['Type'] == 'Sharing':
        msg_content = msg['FileName']    
        # download 并保存在 "文件发送人/文件发送日期"目录下
        # 进一步可以根据 文件类型细分(根据.txt/.mp3等区分)
        file_dir, file_name = msg_from, msg_content
        file_dir = re.sub(rejected_chars, '', file_dir)
        if not os.path.exists(file_dir):
            os.mkdir(file_dir)
        if not os.path.exists("{}/{}".format(file_dir, current_date)):
            os.mkdir("{}/{}".format(file_dir, current_date))
        if msg['Type'] != 'Sharing':
            to_file = "{}/{}/{}".format(file_dir, current_date, file_name)
            msg['Text'](to_file)
        else:
            to_file = "{}/{}/微信分享.xlsx".format(file_dir, current_date)
            try:
                book = openpyxl.load_workbook(to_file)
                sheet = book.get_active_sheet()
            except FileNotFoundError:
                book = openpyxl.Workbook()
                sheet = book.get_active_sheet()
                sheet['A1'] = "主题"
                sheet['B1'] = "链接"
            rows = sheet.max_row
            for row in range(1, rows+1):
                if sheet['B{}'.format(row)].value == msg['Url']:
                    return
            sheet["A{}".format(rows + 1)] = msg['FileName']
            sheet['B{}'.format(rows + 1)] = msg['Url']
            book.save(to_file)
    else:
        return None
    return msg_content
        

@itchat.msg_register([TEXT, PICTURE, FRIENDS, CARD, MAP, SHARING, RECORDING, ATTACHMENT, VIDEO],isFriendChat=True, isGroupChat=True, isMpChat=True)
def handleReceiveMsg(msg):
    global message_id_list
    global message_id
    global friend_list
    global message_list
    # 腾讯企业邮箱的邮件通知
    message_list.append(msg)
    if msg['FileName'] == "新邮件通知":
        if not msg['MsgId'] in message_id_list:
            message_id_list.append(msg['MsgId'])
            print("Email from {}".format(msg['User']['NickName']))
        return None
    if "ActualNickName" not in msg.keys():
        try:
            # 个人消息
            msg_from = friend_list[msg['FromUserName']]
            msg_to = friend_list[msg['ToUserName']]
            message_type = 0
        except KeyError as e:
            # 推送消息
            try:
                msg_from = msg['User']['NickName']
            except:
                msg_from = msg['FromUserName']
            message_type = 1
    else:
        # 群消息
        try:
            msg_from = msg['User']['RemarkName']
            if not len(msg_from):
                msg_from = msg['User']['NickName']
            message_type = 2
        # itchat无法解析有些群的名称
        except KeyError as e:
            return
            
    msg_content = processMsg(msg, msg_from)
    # !!Warning itchat对同一消息会进行多次反馈,需要根据MsgId过滤
    if (not msg_content) or (msg['MsgId'] in message_id_list):
        pass
    else:
        message_id_list.append(msg['MsgId'])
        if message_type == 0:
            msg_content = "ID:{:>3} |{:*^20}|\n{} {} 对 {}: {}".format(message_id, "个人消息", ' '*7, msg_from, msg_to, msg_content)
        elif message_type == 1:
            msg_content = "ID:{:>3} |{:*^20}|\n{} {}: {}".format(message_id, "推送", ' '*7, msg_from, msg_content)
        else:
            msg_content = "ID:{:>3} |{:*^20}|\n{} {}: {}".format(message_id, "群消息From {}".format(msg_from), ' '*7, len(msg['ActualNickName']) and msg['ActualNickName'] or "我", msg_content)
        print (msg_content)
        message_id += 1


if __name__ == "__main__":
    message_id_list = collections.deque(maxlen=5)
    message_list = []
    # 文件夹名不能含有 /\?|*<>:" 
    rejected_chars = r'[:/|><?.*\\"]'
    message_id = 1
    current_date = str(datetime.date.today())
    text_content = None
    itchat.auto_login(hotReload=True)
    friend_list = itchat.get_friends(update=True)
    friend_list = {friend['UserName']:len(friend['RemarkName']) and friend['RemarkName'] or friend['NickName'] for friend in friend_list}
    chatroom_list = itchat.get_chatrooms(update=True)
#    chatroom_list = {chatroom['UserName']:len(chatroom['RemarkName']) and chatroom['RemarkName'] or chatroom['NickName'] for chatroom in chatroom_list}
    itchat.run()
